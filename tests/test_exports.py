# coding: utf-8
import io
import json
import unittest
from collections import OrderedDict
from dateutil import parser
from io import BytesIO, TextIOWrapper
from textwrap import dedent
from zipfile import ZipFile

import openpyxl
import pytest
from path import TempDir

from formpack import FormPack
from formpack.constants import UNTRANSLATED
from formpack.errors import TranslationError
from formpack.schema.fields import (
    ValidationStatusCopyField,
    IdCopyField,
)
from formpack.utils.iterator import get_first_occurrence
from .fixtures import build_fixture, open_fixture_file

customer_satisfaction = build_fixture('customer_satisfaction')
restaurant_profile = build_fixture('restaurant_profile')


class TestFormPackExport(unittest.TestCase):
    maxDiff = None

    def assertTextEqual(self, text1, text2):
        self.assertEqual(dedent(text1).strip(), dedent(text2).strip())

    def test_generator_export(self):

        title, schemas, submissions = customer_satisfaction

        forms = FormPack(schemas, title)
        export = forms.export().to_dict(submissions)
        expected = OrderedDict(
            {
                'Customer Satisfaction': {
                    'fields': ['restaurant_name', 'customer_enjoyment'],
                    'data': [
                        ['Felipes', 'yes'],
                        ['Dunkin Donuts', 'no'],
                        ['McDonalds', 'no'],
                    ],
                }
            }
        )

        self.assertEqual(export, expected)

    def test_generator_export_labels_without_translations(self):

        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)

        self.assertEqual(len(fp[0].translations), 1)

        export = fp.export(lang=UNTRANSLATED).to_dict(submissions)
        expected = OrderedDict(
            {
                'Customer Satisfaction': {
                    'fields': [
                        'Restaurant name',
                        'Did you enjoy your dining experience?',
                    ],
                    'data': [
                        ['Felipes', 'Yes'],
                        ['Dunkin Donuts', 'No'],
                        ['McDonalds', 'No'],
                    ],
                }
            }
        )

        self.assertDictEqual(export, expected)

    def test_generator_export_translation_headers(self):

        title, schemas, submissions = restaurant_profile
        fp = FormPack(schemas, title)

        self.assertEqual(len(fp.versions), 4)
        self.assertEqual(len(fp[1].translations), 2)

        # by default, exports use the question 'name' attribute
        headers = fp.export(versions=0).to_dict(submissions)[
            'Restaurant profile'
        ]['fields']
        self.assertEqual(
            headers,
            [
                'restaurant_name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
            ],
        )

        # the first translation in the list is the translation that
        # appears first in the column list. in this case, 'label::english'
        translations = fp[1].translations
        export = fp.export(lang=translations[0], versions=1)
        data = export.to_dict(submissions)
        headers = data['Restaurant profile']['fields']
        self.assertEqual(
            headers,
            [
                'restaurant name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
            ],
        )

        export = fp.export(lang=translations[1], versions=1)
        data = export.to_dict(submissions)
        headers = data['Restaurant profile']['fields']
        self.assertEqual(
            headers,
            [
                'nom du restaurant',
                'lieu',
                '_lieu_latitude',
                '_lieu_longitude',
                '_lieu_altitude',
                '_lieu_precision',
            ],
        )

        # TODO: make a separate test to test to test __getitem__
        export = fp.export(lang=UNTRANSLATED, versions='rpv1')
        data = export.to_dict(submissions)
        headers = data['Restaurant profile']['fields']
        self.assertEqual(
            headers,
            [
                'restaurant name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
            ],
        )

    def test_export_with_choice_lists(self):

        title, schemas, submissions = restaurant_profile

        fp = FormPack(schemas, title)
        self.assertEqual(len(fp[1].translations), 2)
        # by default, exports use the question 'name' attribute
        options = {'versions': 'rpV3'}

        export = fp.export(**options).to_dict(submissions)['Restaurant profile']
        self.assertEqual(
            export['fields'],
            [
                'restaurant_name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
                'eatery_type',
            ],
        )
        self.assertEqual(
            export['data'],
            [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    '13.42',
                    '-25.43',
                    '',
                    '',
                    'takeaway',
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    '12.43',
                    '-24.53',
                    '',
                    '',
                    'sit_down',
                ],
            ],
        )

        # if a language is passed, fields with available translations
        # are translated into that language
        options['lang'] = fp[1].translations[0]
        export = fp.export(**options).to_dict(submissions)['Restaurant profile']
        self.assertEqual(
            export['data'],
            [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    '13.42',
                    '-25.43',
                    '',
                    '',
                    'take-away',
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    '12.43',
                    '-24.53',
                    '',
                    '',
                    'sit down',
                ],
            ],
        )

        options['lang'] = fp[1].translations[1]
        export = fp.export(**options).to_dict(submissions)['Restaurant profile']
        self.assertEqual(
            export['data'],
            [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    '13.42',
                    '-25.43',
                    '',
                    '',
                    'avec vente à emporter',
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    '12.43',
                    '-24.53',
                    '',
                    '',
                    'traditionnel',
                ],
            ],
        )

    def test_headers_of_group_exports(self):
        title, schemas, submissions = build_fixture('grouped_questions')
        fp = FormPack(schemas, title)
        options = {'versions': 'gqs'}

        # by default, groups are stripped.
        export = fp.export(**options).to_dict(submissions)
        headers = export['Grouped questions']['fields']
        self.assertEqual(
            headers, ['q1', 'g1q1', 'g1sg1q1', 'g1q2', 'g2q1', 'qz']
        )

    def test_headers_of_translated_group_exports(self):
        title, schemas, submissions = build_fixture('grouped_translated')
        fp = FormPack(schemas, title)
        options = {
            'versions': 'grouped_translated_v1',
            'hierarchy_in_labels': True,
        }
        english_export = fp.export(lang='English', **options).to_dict(
            submissions
        )
        self.assertEqual(
            english_export[title]['fields'],
            [
                'start',
                'end',
                'External Characteristics/What kind of symmetry do you have?',
                'External Characteristics/What kind of symmetry do you have?/Spherical',
                'External Characteristics/What kind of symmetry do you have?/Radial',
                'External Characteristics/What kind of symmetry do you have?/Bilateral',
                'External Characteristics/How many segments does your body have?',
                'Do you have body fluids that occupy intracellular space?',
                'Do you descend from an ancestral unicellular organism?',
            ],
        )
        spanish_export = fp.export(lang='Español', **options).to_dict(
            submissions
        )
        self.assertEqual(
            spanish_export[title]['fields'],
            [
                'start',
                'end',
                'Características externas/¿Qué tipo de simetría tiene?',
                'Características externas/¿Qué tipo de simetría tiene?/Esférico',
                'Características externas/¿Qué tipo de simetría tiene?/Radial',
                'Características externas/¿Qué tipo de simetría tiene?/Bilateral',
                'Características externas/¿Cuántos segmentos tiene tu cuerpo?',
                '¿Tienes fluidos corporales que ocupan espacio intracelular?',
                '¿Desciende de un organismo unicelular ancestral?',
            ],
        )

    def assertDictEquals(self, arg1, arg2):
        _j = lambda arg: json.dumps(arg, indent=4, sort_keys=True)
        assert _j(arg1) == _j(arg2)

    def test_submissions_of_group_exports(self):
        title, schemas, submissions = build_fixture('grouped_questions')
        fp = FormPack(schemas, title)
        options = {'versions': 'gqs'}

        export = fp.export(**options).to_dict(submissions)['Grouped questions']
        self.assertDictEquals(
            export['fields'], ['q1', 'g1q1', 'g1sg1q1', 'g1q2', 'g2q1', 'qz']
        )
        self.assertDictEquals(
            export['data'],
            [
                [
                    'respondent1\'s r1',
                    'respondent1\'s r2',
                    'respondent1\'s r2.5',
                    'respondent1\'s r2.75 :)',
                    'respondent1\'s r3',
                    'respondent1\'s r4',
                ],
                [
                    'respondent2\'s r1',
                    'respondent2\'s r2',
                    'respondent2\'s r2.5',
                    'respondent2\'s r2.75 :)',
                    'respondent2\'s r3',
                    'respondent2\'s r4',
                ],
            ],
        )

        options['hierarchy_in_labels'] = '/'
        export = fp.export(**options).to_dict(submissions)['Grouped questions']
        self.assertDictEquals(
            export['fields'],
            ['q1', 'g1/g1q1', 'g1/sg1/g1sg1q1', 'g1/g1q2', 'g2/g2q1', 'qz'],
        )
        self.assertDictEquals(
            export['data'],
            [
                [
                    'respondent1\'s r1',
                    'respondent1\'s r2',
                    'respondent1\'s r2.5',
                    'respondent1\'s r2.75 :)',
                    'respondent1\'s r3',
                    'respondent1\'s r4',
                ],
                [
                    'respondent2\'s r1',
                    'respondent2\'s r2',
                    'respondent2\'s r2.5',
                    'respondent2\'s r2.75 :)',
                    'respondent2\'s r3',
                    'respondent2\'s r4',
                ],
            ],
        )

    def test_translations_labels_mismatch(self):
        title, schemas, submissions = build_fixture(
            'translations_labels_mismatch'
        )
        with self.assertRaises(TranslationError):
            FormPack(schemas, title)

    def test_simple_nested_grouped_repeatable(self):
        title, schemas, submissions = build_fixture(
            'simple_nested_grouped_repeatable'
        )
        fp = FormPack(schemas, title)
        options = {'versions': fp.versions}
        export = fp.export(**options)
        actual_dict = export.to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Simple nested grouped repeatable',
                    {'data': [[1], [2]], 'fields': ['_index']},
                ),
                (
                    'cities',
                    {
                        'data': [
                            [1, 'Simple nested grouped repeatable', 1],
                            [2, 'Simple nested grouped repeatable', 2],
                        ],
                        'fields': [
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                    },
                ),
                (
                    'respondents',
                    {
                        'data': [
                            ['Caesar', '', 'cities', 1],
                            ['Augustus', '', 'cities', 1],
                            ['Caesar', '55', 'cities', 2],
                            ['Augustus', '75', 'cities', 2],
                        ],
                        'fields': [
                            'respondent_name',
                            'respondent_age',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                    },
                ),
                (
                    'items',
                    {
                        'data': [
                            ['Sword', 'cities', 2],
                            ['Thrown', 'cities', 2],
                        ],
                        'fields': [
                            'item',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                    },
                ),
            ]
        )

        assert 4 == len(actual_dict)
        assert expected_dict == actual_dict

        with TempDir() as d:
            xls = d / 'foo.xlsx'
            export.to_xlsx(xls, submissions)
            assert xls.isfile()

    def test_repeats(self):
        title, schemas, submissions = build_fixture('grouped_repeatable')
        fp = FormPack(schemas, title)
        options = {'versions': 'rgv1'}
        export = fp.export(**options).to_dict(submissions)
        self.assertEqual(
            export,
            OrderedDict(
                [
                    (
                        'Household survey with repeatable groups',
                        {
                            'fields': [
                                'start',
                                'end',
                                'household_location',
                                '_index',
                            ],
                            'data': [
                                [
                                    '2016-03-14T14:15:48.000-04:00',
                                    '2016-03-14T14:18:35.000-04:00',
                                    'montreal',
                                    1,
                                ],
                                [
                                    '2016-03-14T14:14:10.000-04:00',
                                    '2016-03-14T14:15:48.000-04:00',
                                    'marseille',
                                    2,
                                ],
                                [
                                    '2016-03-14T14:13:53.000-04:00',
                                    '2016-03-14T14:14:10.000-04:00',
                                    'rocky mountains',
                                    3,
                                ],
                                [
                                    '2016-03-14T14:12:54.000-04:00',
                                    '2016-03-14T14:13:53.000-04:00',
                                    'toronto',
                                    4,
                                ],
                                [
                                    '2016-03-14T14:18:35.000-04:00',
                                    '2016-03-14T15:19:20.000-04:00',
                                    'new york',
                                    5,
                                ],
                                [
                                    '2016-03-14T14:11:25.000-04:00',
                                    '2016-03-14T14:12:03.000-04:00',
                                    'boston',
                                    6,
                                ],
                            ],
                        },
                    ),
                    (
                        'houshold_member_repeat',
                        {
                            'fields': [
                                'household_member_name',
                                '_parent_table_name',
                                '_parent_index',
                            ],
                            'data': [
                                [
                                    'peter',
                                    'Household survey with repeatable groups',
                                    1,
                                ],
                                [
                                    'kyle',
                                    'Household survey with repeatable groups',
                                    2,
                                ],
                                [
                                    'linda',
                                    'Household survey with repeatable groups',
                                    2,
                                ],
                                [
                                    'morty',
                                    'Household survey with repeatable groups',
                                    3,
                                ],
                                [
                                    'tony',
                                    'Household survey with repeatable groups',
                                    4,
                                ],
                                [
                                    'mary',
                                    'Household survey with repeatable groups',
                                    4,
                                ],
                                [
                                    'emma',
                                    'Household survey with repeatable groups',
                                    5,
                                ],
                                [
                                    'parker',
                                    'Household survey with repeatable groups',
                                    5,
                                ],
                                [
                                    'amadou',
                                    'Household survey with repeatable groups',
                                    6,
                                ],
                                [
                                    'esteban',
                                    'Household survey with repeatable groups',
                                    6,
                                ],
                                [
                                    'suzie',
                                    'Household survey with repeatable groups',
                                    6,
                                ],
                                [
                                    'fiona',
                                    'Household survey with repeatable groups',
                                    6,
                                ],
                                [
                                    'phillip',
                                    'Household survey with repeatable groups',
                                    6,
                                ],
                            ],
                        },
                    ),
                ]
            ),
        )

    def test_select_one_legacy(self):
        title, schemas, submissions = build_fixture('select_one_legacy')
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1'}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Your favourite Roman emperors',
                    {
                        'fields': [
                            'fav_emperor',
                        ],
                        'data': [
                            [
                                'julius',
                            ],
                            [
                                'augustus',
                            ],
                            [
                                'tiberius',
                            ],
                        ],
                    },
                )
            ]
        )
        self.assertEqual(export, expected_dict)

    def test_media_types_include_media_url(self):
        # need to make sure that filenames such as "another-julius 1).jpg"
        # don't break the export
        title, schemas, submissions = build_fixture('media_types')
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1', 'include_media_url': True}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Media of your favourite Roman emperors',
                    {
                        'fields': [
                            'audit',
                            'audit_URL',
                            'fav_emperor',
                            'image_of_emperor',
                            'image_of_emperor_URL',
                            'another_image_of_emperor',
                            'another_image_of_emperor_URL',
                        ],
                        'data': [
                            [
                                'audit.csv',
                                'https://kc.kobo.org/media/original?media_file=/path/to/audit.csv',
                                'julius',
                                'julius.jpg',
                                'https://kc.kobo.org/media/original?media_file=/path/to/julius.jpg',
                                'another-julius 1).jpg',
                                'https://kc.kobo.org/media/original?media_file=/path/to/another-julius_1.jpg',
                            ],
                            [
                                'audit.csv',
                                'https://kc.kobo.org/media/original?media_file=/path/to/audit.csv',
                                'augustus',
                                'augustus.jpg',
                                'https://kc.kobo.org/media/original?media_file=/path/to/augustus.jpg',
                                '',
                                '',
                            ],
                        ],
                    },
                )
            ]
        )
        assert export == expected_dict

    def test_media_types_exclude_media_url(self):
        title, schemas, submissions = build_fixture('media_types')
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1'}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Media of your favourite Roman emperors',
                    {
                        'fields': [
                            'audit',
                            'fav_emperor',
                            'image_of_emperor',
                            'another_image_of_emperor',
                        ],
                        'data': [
                            [
                                'audit.csv',
                                'julius',
                                'julius.jpg',
                                'another-julius 1).jpg',
                            ],
                            [
                                'audit.csv',
                                'augustus',
                                'augustus.jpg',
                                '',
                            ],
                        ],
                    },
                )
            ]
        )
        assert export == expected_dict

    def test_select_one_from_previous_answers(self):
        title, schemas, submissions = build_fixture(
            'select_one_from_previous_answers'
        )
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1'}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Household survey with select_one from previous answers',
                    {
                        'fields': [
                            'Q1',
                            'Q4',
                            'Q5',
                            '_index',
                        ],
                        'data': [
                            [
                                '2',
                                'Julius Caesar',
                                'Gaius Octavius',
                                1,
                            ]
                        ],
                    },
                ),
                (
                    'FM',
                    {
                        'fields': [
                            'Q2',
                            'Q3',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            [
                                'Julius Caesar',
                                '53',
                                'Household survey with select_one from previous answers',
                                1,
                            ],
                            [
                                'Gaius Octavius',
                                '17',
                                'Household survey with select_one from previous answers',
                                1,
                            ],
                        ],
                    },
                ),
            ]
        )
        self.assertEqual(export, expected_dict)

    def test_select_one_from_previous_answers_xls_types(self):
        title, schemas, submissions = build_fixture(
            'select_one_from_previous_answers'
        )
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1', 'xls_types_as_text': False}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Household survey with select_one from previous answers',
                    {
                        'fields': [
                            'Q1',
                            'Q4',
                            'Q5',
                            '_index',
                        ],
                        'data': [
                            [
                                2,
                                'Julius Caesar',
                                'Gaius Octavius',
                                1,
                            ]
                        ],
                    },
                ),
                (
                    'FM',
                    {
                        'fields': [
                            'Q2',
                            'Q3',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            [
                                'Julius Caesar',
                                53,
                                'Household survey with select_one from previous answers',
                                1,
                            ],
                            [
                                'Gaius Octavius',
                                17,
                                'Household survey with select_one from previous answers',
                                1,
                            ],
                        ],
                    },
                ),
            ]
        )
        self.assertEqual(export, expected_dict)

    def test_select_or_other(self):
        title, schemas, submissions = build_fixture('or_other')
        fp = FormPack(schemas, title)
        options = {'versions': 'romev1'}
        export = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Your favourite Roman emperors or other',
                    {
                        'fields': [
                            'fav_emperor',
                            'fav_emperor_other',
                            'fav_emperors',
                            'fav_emperors/julius',
                            'fav_emperors/augustus',
                            'fav_emperors/tiberius',
                            'fav_emperors/caligula',
                            'fav_emperors/other',
                            'fav_emperors_other',
                        ],
                        'data': [
                            [
                                'other',
                                'Nero',
                                'caligula other',
                                '0',
                                '0',
                                '0',
                                '1',
                                '1',
                                'Commodus',
                            ],
                            [
                                'augustus',
                                '',
                                'julius tiberius',
                                '1',
                                '0',
                                '1',
                                '0',
                                '0',
                                '',
                            ],
                            [
                                'other',
                                'Marcus Aurelius',
                                'julius augustus',
                                '1',
                                '1',
                                '0',
                                '0',
                                '0',
                                '',
                            ],
                        ],
                    },
                )
            ]
        )
        self.assertEqual(export, expected_dict)

    def test_nested_repeats_with_copy_fields(self):
        title, schemas, submissions = build_fixture('nested_grouped_repeatable')
        fp = FormPack(schemas, title)
        export_dict = fp.export(
            versions='bird_nests_v1',
            copy_fields=(IdCopyField, '_uuid', ValidationStatusCopyField),
        ).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Bird nest survey with nested repeatable groups',
                    {
                        'fields': [
                            'start',
                            'end',
                            '_id',
                            '_uuid',
                            '_validation_status',
                            '_index',
                        ],
                        'data': [
                            [
                                '2017-12-27T15:53:26.000-05:00',
                                '2017-12-27T15:58:20.000-05:00',
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                                1,
                            ],
                            [
                                '2017-12-27T15:58:20.000-05:00',
                                '2017-12-27T15:58:50.000-05:00',
                                124,
                                '790af158-7b24-4651-b584-27bf65b9e397',
                                'validation_status_not_approved',
                                2,
                            ],
                        ],
                    },
                ),
                (
                    'group_tree',
                    {
                        'fields': [
                            'What_kind_of_tree_is_this',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                            '_submission__id',
                            '_submission__uuid',
                            '_submission__validation_status',
                        ],
                        'data': [
                            [
                                'pine',
                                1,
                                'Bird nest survey with nested repeatable groups',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'spruce',
                                2,
                                'Bird nest survey with nested repeatable groups',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'maple',
                                3,
                                'Bird nest survey with nested repeatable groups',
                                2,
                                124,
                                '790af158-7b24-4651-b584-27bf65b9e397',
                                'validation_status_not_approved',
                            ],
                        ],
                    },
                ),
                (
                    'group_nest',
                    {
                        'fields': [
                            'How_high_above_the_ground_is_the_nest',
                            'How_many_eggs_are_in_the_nest',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                            '_submission__id',
                            '_submission__uuid',
                            '_submission__validation_status',
                        ],
                        'data': [
                            [
                                '13',
                                '3',
                                1,
                                'group_tree',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                '15',
                                '1',
                                2,
                                'group_tree',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                '10',
                                '2',
                                3,
                                'group_tree',
                                2,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                '23',
                                '1',
                                4,
                                'group_tree',
                                3,
                                124,
                                '790af158-7b24-4651-b584-27bf65b9e397',
                                'validation_status_not_approved',
                            ],
                        ],
                    },
                ),
                (
                    'group_egg',
                    {
                        'fields': [
                            'Describe_the_egg',
                            '_parent_table_name',
                            '_parent_index',
                            '_submission__id',
                            '_submission__uuid',
                            '_submission__validation_status',
                        ],
                        'data': [
                            [
                                'brown and speckled; medium',
                                'group_nest',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'brown and speckled; large; cracked',
                                'group_nest',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'light tan; small',
                                'group_nest',
                                1,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'cream-colored',
                                'group_nest',
                                2,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'reddish-brown; medium',
                                'group_nest',
                                3,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'reddish-brown; small',
                                'group_nest',
                                3,
                                123,
                                'f16d9a3f-0892-413e-81d4-758ab188ea0b',
                                'validation_status_approved',
                            ],
                            [
                                'grey and speckled',
                                'group_nest',
                                4,
                                124,
                                '790af158-7b24-4651-b584-27bf65b9e397',
                                'validation_status_not_approved',
                            ],
                        ],
                    },
                ),
            ]
        )
        self.assertEqual(export_dict, expected_dict)

    def test_nested_repeats(self):
        title, schemas, submissions = build_fixture('nested_grouped_repeatable')
        fp = FormPack(schemas, title)
        export_dict = fp.export(versions='bird_nests_v1').to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Bird nest survey with nested repeatable groups',
                    {
                        'fields': ['start', 'end', '_index'],
                        'data': [
                            [
                                '2017-12-27T15:53:26.000-05:00',
                                '2017-12-27T15:58:20.000-05:00',
                                1,
                            ],
                            [
                                '2017-12-27T15:58:20.000-05:00',
                                '2017-12-27T15:58:50.000-05:00',
                                2,
                            ],
                        ],
                    },
                ),
                (
                    'group_tree',
                    {
                        'fields': [
                            'What_kind_of_tree_is_this',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            [
                                'pine',
                                1,
                                'Bird nest survey with nested repeatable groups',
                                1,
                            ],
                            [
                                'spruce',
                                2,
                                'Bird nest survey with nested repeatable groups',
                                1,
                            ],
                            [
                                'maple',
                                3,
                                'Bird nest survey with nested repeatable groups',
                                2,
                            ],
                        ],
                    },
                ),
                (
                    'group_nest',
                    {
                        'fields': [
                            'How_high_above_the_ground_is_the_nest',
                            'How_many_eggs_are_in_the_nest',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            ['13', '3', 1, 'group_tree', 1],
                            ['15', '1', 2, 'group_tree', 1],
                            ['10', '2', 3, 'group_tree', 2],
                            ['23', '1', 4, 'group_tree', 3],
                        ],
                    },
                ),
                (
                    'group_egg',
                    {
                        'fields': [
                            'Describe_the_egg',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            ['brown and speckled; medium', 'group_nest', 1],
                            [
                                'brown and speckled; large; cracked',
                                'group_nest',
                                1,
                            ],
                            ['light tan; small', 'group_nest', 1],
                            ['cream-colored', 'group_nest', 2],
                            ['reddish-brown; medium', 'group_nest', 3],
                            ['reddish-brown; small', 'group_nest', 3],
                            ['grey and speckled', 'group_nest', 4],
                        ],
                    },
                ),
            ]
        )
        self.assertEqual(export_dict, expected_dict)

    def test_nested_repeats_with_xls_types(self):
        title, schemas, submissions = build_fixture('nested_grouped_repeatable')
        fp = FormPack(schemas, title)
        options = {'versions': 'bird_nests_v2', 'xls_types_as_text': False}
        export_dict = fp.export(**options).to_dict(submissions)
        expected_dict = OrderedDict(
            [
                (
                    'Bird nest survey with nested repeatable groups',
                    {
                        'fields': ['start', 'end', '_index'],
                        'data': [
                            [
                                parser.parse('2017-12-27T15:53:26.000-05:00'),
                                parser.parse('2017-12-27T15:58:20.000-05:00'),
                                1,
                            ],
                            [
                                parser.parse('2017-12-27T15:58:20.000-05:00'),
                                parser.parse('2017-12-27T15:58:50.000-05:00'),
                                2,
                            ],
                        ],
                    },
                ),
                (
                    'group_tree',
                    {
                        'fields': [
                            'What_kind_of_tree_is_this',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            [
                                'pine',
                                1,
                                'Bird nest survey with nested repeatable groups',
                                1,
                            ],
                            [
                                'spruce',
                                2,
                                'Bird nest survey with nested repeatable groups',
                                1,
                            ],
                            [
                                'nan',
                                3,
                                'Bird nest survey with nested repeatable groups',
                                2,
                            ],
                        ],
                    },
                ),
                (
                    'group_nest',
                    {
                        'fields': [
                            'How_high_above_the_ground_is_the_nest',
                            'How_many_eggs_are_in_the_nest',
                            '_index',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            [13, 3, 1, 'group_tree', 1],
                            [15, 1, 2, 'group_tree', 1],
                            [10, 2, 3, 'group_tree', 2],
                            [23, 1, 4, 'group_tree', 3],
                        ],
                    },
                ),
                (
                    'group_egg',
                    {
                        'fields': [
                            'Describe_the_egg',
                            '_parent_table_name',
                            '_parent_index',
                        ],
                        'data': [
                            ['brown and speckled; medium', 'group_nest', 1],
                            [
                                'brown and speckled; large; cracked',
                                'group_nest',
                                1,
                            ],
                            ['light tan; small', 'group_nest', 1],
                            ['cream-colored', 'group_nest', 2],
                            ['reddish-brown; medium', 'group_nest', 3],
                            ['reddish-brown; small', 'group_nest', 3],
                            ['grey and speckled', 'group_nest', 4],
                        ],
                    },
                ),
            ]
        )
        self.assertEqual(export_dict, expected_dict)

    def test_repeats_alias(self):
        title, schemas, submissions = build_fixture('grouped_repeatable_alias')
        fp = FormPack(schemas, title)
        options = {'versions': 'rgv1'}
        export = fp.export(**options).to_dict(submissions)

        self.assertEqual(
            export,
            OrderedDict(
                [
                    (
                        'Grouped Repeatable Alias',
                        {
                            'fields': [
                                'start',
                                'end',
                                'household_location',
                                '_index',
                            ],
                            'data': [
                                [
                                    '2016-03-14T14:15:48.000-04:00',
                                    '2016-03-14T14:18:35.000-04:00',
                                    'montreal',
                                    1,
                                ],
                                [
                                    '2016-03-14T14:14:10.000-04:00',
                                    '2016-03-14T14:15:48.000-04:00',
                                    'marseille',
                                    2,
                                ],
                                [
                                    '2016-03-14T14:13:53.000-04:00',
                                    '2016-03-14T14:14:10.000-04:00',
                                    'rocky mountains',
                                    3,
                                ],
                                [
                                    '2016-03-14T14:12:54.000-04:00',
                                    '2016-03-14T14:13:53.000-04:00',
                                    'toronto',
                                    4,
                                ],
                                [
                                    '2016-03-14T14:18:35.000-04:00',
                                    '2016-03-14T15:19:20.000-04:00',
                                    'new york',
                                    5,
                                ],
                                [
                                    '2016-03-14T14:11:25.000-04:00',
                                    '2016-03-14T14:12:03.000-04:00',
                                    'boston',
                                    6,
                                ],
                            ],
                        },
                    ),
                    (
                        'houshold_member_repeat',
                        {
                            'fields': [
                                'household_member_name',
                                '_parent_table_name',
                                '_parent_index',
                            ],
                            'data': [
                                ['peter', 'Grouped Repeatable Alias', 1],
                                ['kyle', 'Grouped Repeatable Alias', 2],
                                ['linda', 'Grouped Repeatable Alias', 2],
                                ['morty', 'Grouped Repeatable Alias', 3],
                                ['tony', 'Grouped Repeatable Alias', 4],
                                ['mary', 'Grouped Repeatable Alias', 4],
                                ['emma', 'Grouped Repeatable Alias', 5],
                                ['parker', 'Grouped Repeatable Alias', 5],
                                ['amadou', 'Grouped Repeatable Alias', 6],
                                ['esteban', 'Grouped Repeatable Alias', 6],
                                ['suzie', 'Grouped Repeatable Alias', 6],
                                ['fiona', 'Grouped Repeatable Alias', 6],
                                ['phillip', 'Grouped Repeatable Alias', 6],
                            ],
                        },
                    ),
                ]
            ),
        )

    def test_substitute_xml_names_for_missing_labels(self):
        title, schemas, submissions = build_fixture('grouped_translated')

        # Remove a choice's labels
        self.assertEqual(
            schemas[0]['content']['choices'][0]['label'],
            ['Spherical', 'Esférico'],
        )
        del schemas[0]['content']['choices'][0]['label']

        # Remove a group's labels
        self.assertEqual(
            schemas[0]['content']['survey'][2]['label'],
            ['External Characteristics', 'Características externas'],
        )
        del schemas[0]['content']['survey'][2]['label']

        # Remove a grouped question's labels
        self.assertEqual(
            schemas[0]['content']['survey'][4]['label'],
            [
                'How many segments does your body have?',
                '¿Cuántos segmentos tiene tu cuerpo?',
            ],
        )
        del schemas[0]['content']['survey'][4]['label']

        # Remove a non-grouped question's labels
        self.assertEqual(
            schemas[0]['content']['survey'][6]['label'],
            [
                'Do you have body fluids that occupy intracellular space?',
                '¿Tienes fluidos corporales que ocupan espacio intracelular?',
            ],
        )
        del schemas[0]['content']['survey'][6]['label']

        fp = FormPack(schemas, title)
        options = {
            'versions': 'grouped_translated_v1',
            'hierarchy_in_labels': True,
        }

        # Missing labels should be replaced with XML names
        english_export = fp.export(lang='English', **options).to_dict(
            submissions
        )
        self.assertEqual(
            english_export[title]['fields'],
            [
                'start',
                'end',
                'external_characteristics/What kind of symmetry do you have?',
                'external_characteristics/What kind of symmetry do you have?/spherical',
                'external_characteristics/What kind of symmetry do you have?/Radial',
                'external_characteristics/What kind of symmetry do you have?/Bilateral',
                'external_characteristics/How_many_segments_does_your_body_have',
                'Do_you_have_body_flu_intracellular_space',
                'Do you descend from an ancestral unicellular organism?',
            ],
        )
        self.assertEqual(
            english_export[title]['data'][0][2], 'spherical Radial Bilateral'
        )
        spanish_export = fp.export(lang='Español', **options).to_dict(
            submissions
        )
        self.assertEqual(
            spanish_export[title]['fields'],
            [
                'start',
                'end',
                'external_characteristics/¿Qué tipo de simetría tiene?',
                'external_characteristics/¿Qué tipo de simetría tiene?/spherical',
                'external_characteristics/¿Qué tipo de simetría tiene?/Radial',
                'external_characteristics/¿Qué tipo de simetría tiene?/Bilateral',
                'external_characteristics/How_many_segments_does_your_body_have',
                'Do_you_have_body_flu_intracellular_space',
                '¿Desciende de un organismo unicelular ancestral?',
            ],
        )
        self.assertEqual(
            spanish_export[title]['data'][0][2], 'spherical Radial Bilateral'
        )

    def test_substitute_xml_names_for_missing_translations(self):
        title, schemas, submissions = build_fixture('grouped_translated')

        # Remove a choice's translation
        self.assertEqual(
            schemas[0]['content']['choices'][0]['label'],
            ['Spherical', 'Esférico'],
        )
        schemas[0]['content']['choices'][0]['label'] = [
            'Spherical',
            UNTRANSLATED,
        ]

        # Remove a group's translation
        self.assertEqual(
            schemas[0]['content']['survey'][2]['label'],
            ['External Characteristics', 'Características externas'],
        )
        schemas[0]['content']['survey'][2]['label'] = [
            'External Characteristics',
            UNTRANSLATED,
        ]

        # Remove a grouped question's translation
        self.assertEqual(
            schemas[0]['content']['survey'][4]['label'],
            [
                'How many segments does your body have?',
                '¿Cuántos segmentos tiene tu cuerpo?',
            ],
        )
        schemas[0]['content']['survey'][4]['label'] = [
            'How many segments does your body have?',
            UNTRANSLATED,
        ]

        # Remove a non-grouped question's translation
        self.assertEqual(
            schemas[0]['content']['survey'][6]['label'],
            [
                'Do you have body fluids that occupy intracellular space?',
                '¿Tienes fluidos corporales que ocupan espacio intracelular?',
            ],
        )
        schemas[0]['content']['survey'][6]['label'] = [
            'Do you have body fluids that occupy intracellular space?',
            UNTRANSLATED,
        ]

        fp = FormPack(schemas, title)
        options = {
            'versions': 'grouped_translated_v1',
            'hierarchy_in_labels': True,
        }

        # All the English translations should still be present
        english_export = fp.export(lang='English', **options).to_dict(
            submissions
        )
        self.assertEqual(
            english_export[title]['fields'],
            [
                'start',
                'end',
                'External Characteristics/What kind of symmetry do you have?',
                'External Characteristics/What kind of symmetry do you have?/Spherical',
                'External Characteristics/What kind of symmetry do you have?/Radial',
                'External Characteristics/What kind of symmetry do you have?/Bilateral',
                'External Characteristics/How many segments does your body have?',
                'Do you have body fluids that occupy intracellular space?',
                'Do you descend from an ancestral unicellular organism?',
            ],
        )
        self.assertEqual(
            english_export[title]['data'][0][2], 'Spherical Radial Bilateral'
        )

        # Missing Spanish translations should be replaced with XML names
        spanish_export = fp.export(lang='Español', **options).to_dict(
            submissions
        )
        self.assertEqual(
            spanish_export[title]['fields'],
            [
                'start',
                'end',
                'external_characteristics/¿Qué tipo de simetría tiene?',
                'external_characteristics/¿Qué tipo de simetría tiene?/spherical',
                'external_characteristics/¿Qué tipo de simetría tiene?/Radial',
                'external_characteristics/¿Qué tipo de simetría tiene?/Bilateral',
                'external_characteristics/How_many_segments_does_your_body_have',
                'Do_you_have_body_flu_intracellular_space',
                '¿Desciende de un organismo unicelular ancestral?',
            ],
        )
        self.assertEqual(
            spanish_export[title]['data'][0][2], 'spherical Radial Bilateral'
        )

    def test_csv(self):
        title, schemas, submissions = build_fixture('grouped_questions')
        fp = FormPack(schemas, title)
        options = {'versions': 'gqs'}
        csv_data = '\n'.join(fp.export(**options).to_csv(submissions))

        expected = """
        "q1";"g1q1";"g1sg1q1";"g1q2";"g2q1";"qz"
        "respondent1's r1";"respondent1's r2";"respondent1's r2.5";"respondent1's r2.75 :)";"respondent1's r3";"respondent1's r4"
        "respondent2's r1";"respondent2's r2";"respondent2's r2.5";"respondent2's r2.75 :)";"respondent2's r3";"respondent2's r4"
        """

        self.assertTextEqual(csv_data, expected)

        options = {'versions': 'gqs', 'hierarchy_in_labels': True}
        csv_data = '\n'.join(fp.export(**options).to_csv(submissions))

        expected = """
        "q1";"g1/g1q1";"g1/sg1/g1sg1q1";"g1/g1q2";"g2/g2q1";"qz"
        "respondent1's r1";"respondent1's r2";"respondent1's r2.5";"respondent1's r2.75 :)";"respondent1's r3";"respondent1's r4"
        "respondent2's r1";"respondent2's r2";"respondent2's r2.5";"respondent2's r2.75 :)";"respondent2's r3";"respondent2's r4"
        """

        self.assertTextEqual(csv_data, expected)

        options = {
            'versions': 'gqs',
            'hierarchy_in_labels': True,
            'lang': UNTRANSLATED,
        }
        csv_data = '\n'.join(fp.export(**options).to_csv(submissions))

        expected = """
        "Q1";"Group 1/G1Q1";"Group 1/Sub Group 1/G1SG1Q1";"Group 1/G1Q2";"g2/G2Q1";"QZed"
        "respondent1's r1";"respondent1's r2";"respondent1's r2.5";"respondent1's r2.75 :)";"respondent1's r3";"respondent1's r4"
        "respondent2's r1";"respondent2's r2";"respondent2's r2.5";"respondent2's r2.75 :)";"respondent2's r3";"respondent2's r4"
        """
        self.assertTextEqual(csv_data, expected)

        title, schemas, submissions = restaurant_profile
        fp = FormPack(schemas, title)
        options = {'versions': 'rpV3', 'lang': fp[1].translations[1]}
        csv_data = '\n'.join(fp.export(**options).to_csv(submissions))

        expected = """
        "nom du restaurant";"lieu";"_lieu_latitude";"_lieu_longitude";"_lieu_altitude";"_lieu_precision";"type de restaurant"
        "Taco Truck";"13.42 -25.43";"13.42";"-25.43";"";"";"avec vente à emporter"
        "Harvest";"12.43 -24.53";"12.43";"-24.53";"";"";"traditionnel"
        """

        self.assertTextEqual(csv_data, expected)

    def test_csv_quote_escaping(self):
        title, schemas, submissions = build_fixture(
            'quotes_newlines_and_long_urls'
        )
        fp = FormPack(schemas, title)
        csv_lines = list(fp.export().to_csv(submissions))
        expected_lines = []
        expected_lines.append(
            '"Enter_some_long_text_and_linebreaks_here";'
            '"Some_other_question"'
        )
        expected_lines.append(
            '"Check out this URL I found:\nhttps://now.read.this/?Never%20forg'
            'et%20that%20you%20are%20one%20of%20a%20kind.%20Never%20forget%20t'
            'hat%20if%20there%20weren%27t%20any%20need%20for%20you%20in%20all%'
            '20your%20uniqueness%20to%20be%20on%20this%20earth%2C%20you%20woul'
            'dn%27t%20be%20here%20in%20the%20first%20place.%20And%20never%20fo'
            'rget%2C%20no%20matter%20how%20overwhelming%20life%27s%20challenge'
            's%20and%20problems%20seem%20to%20be%2C%20that%20one%20person%20ca'
            'n%20make%20a%20difference%20in%20the%20world.%20In%20fact%2C%20it'
            '%20is%20always%20because%20of%20one%20person%20that%20all%20the%2'
            '0changes%20that%20matter%20in%20the%20world%20come%20about.%20So%'
            '20be%20that%20one%20person' + ('!' * 3000) + '";"yes"'
        )
        expected_lines.append(
            '"Hi, my name is Roger.""\n\nI like to enter quotes randomly and '
            'follow them with new lines.";"yes"'
        )
        expected_lines.append('"This one has no linebreaks";"no"')
        expected_lines.append('"This\nis\nnot\na Haiku";"yes"')
        expected_lines.append(
            '"""Hands up!"" He yelled.\nWhy?""\n'
            '''She couldn't understand anything.";"yes"'''
        )

        self.assertListEqual(csv_lines, expected_lines)

    def test_csv_with_tag_headers(self):
        title, schemas, submissions = build_fixture('dietary_needs')
        fp = FormPack(schemas, title)
        options = {'versions': 'dietv1', 'tag_cols_for_header': ['hxl']}
        rows = list(fp.export(**options).to_csv(submissions))
        assert rows[1] == ('"#loc+name";"#indicator+diet";"";"";"";""')

    def test_csv_with_tag_headers_select_multiple_summary_or_details(self):
        """
        The tag header row needs to change in accordance with the
        `multiple_select` export option
        """
        title, schemas, submissions = build_fixture('dietary_needs')
        fp = FormPack(schemas, title)
        options = {'versions': 'dietv1', 'tag_cols_for_header': ['hxl']}

        rows = list(
            fp.export(multiple_select='summary', **options).to_csv(submissions)
        )
        assert rows[1] == ('"#loc+name";"#indicator+diet"')

        rows = list(
            fp.export(multiple_select='details', **options).to_csv(submissions)
        )
        assert rows[1] == ('"#loc+name";"#indicator+diet";"";"";""')

    # disabled for now
    # @raises(RuntimeError)
    # def test_csv_on_repeatable_groups(self):
    #     title, schemas, submissions = build_fixture('grouped_repeatable')
    #     fp = FormPack(schemas, title)
    #     options = {'versions': 'rgv1'}
    #     list(fp.export(**options).to_csv(submissions))

    def test_export_with_split_fields(self):
        title, schemas, submissions = restaurant_profile
        fp = FormPack(schemas, title)
        options = {'versions': 'rpV4'}
        export = fp.export(**options).to_dict(submissions)['Restaurant profile']
        expected = {
            'fields': [
                'restaurant_name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
                'eatery_type',
                'eatery_type/sit_down',
                'eatery_type/takeaway',
            ],
            'data': [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    '13.42',
                    '-25.43',
                    '',
                    '',
                    'takeaway sit_down',
                    '1',
                    '1',
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    '12.43',
                    '-24.53',
                    '',
                    '',
                    'sit_down',
                    '1',
                    '0',
                ],
                [
                    'Wololo',
                    '12.43 -24.54 1 0',
                    '12.43',
                    '-24.54',
                    '1',
                    '0',
                    '',
                    '0',
                    '0',
                ],
                [
                    'Los pollos hermanos',
                    '12.43 -24.54 1',
                    '12.43',
                    '-24.54',
                    '1',
                    '',
                    '',
                    '',
                    '',
                ],
            ],
        }

        self.assertEqual(export, expected)

        options = {
            'versions': 'rpV4',
            'group_sep': '::',
            'hierarchy_in_labels': True,
            'lang': fp[-1].translations[1],
        }
        export = fp.export(**options).to_dict(submissions)['Restaurant profile']

        expected = {
            'fields': [
                'nom du restaurant',
                'lieu',
                '_lieu_latitude',
                '_lieu_longitude',
                '_lieu_altitude',
                '_lieu_precision',
                'type de restaurant',
                'type de restaurant::traditionnel',
                'type de restaurant::avec vente à emporter',
            ],
            'data': [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    '13.42',
                    '-25.43',
                    '',
                    '',
                    'avec vente à emporter traditionnel',
                    '1',
                    '1',
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    '12.43',
                    '-24.53',
                    '',
                    '',
                    'traditionnel',
                    '1',
                    '0',
                ],
                [
                    'Wololo',
                    '12.43 -24.54 1 0',
                    '12.43',
                    '-24.54',
                    '1',
                    '0',
                    '',
                    '0',
                    '0',
                ],
                [
                    'Los pollos hermanos',
                    '12.43 -24.54 1',
                    '12.43',
                    '-24.54',
                    '1',
                    '',
                    '',
                    '',
                    '',
                ],
            ],
        }

        self.assertEqual(export, expected)

    def test_export_with_split_fields_gps_fields_and_multiple_selects_xls_types(
        self,
    ):
        title, schemas, submissions = restaurant_profile
        fp = FormPack(schemas, title)
        options = {'versions': 'rpV4', 'xls_types_as_text': False}
        export = fp.export(**options).to_dict(submissions)['Restaurant profile']
        expected = {
            'fields': [
                'restaurant_name',
                'location',
                '_location_latitude',
                '_location_longitude',
                '_location_altitude',
                '_location_precision',
                'eatery_type',
                'eatery_type/sit_down',
                'eatery_type/takeaway',
            ],
            'data': [
                [
                    'Taco Truck',
                    '13.42 -25.43',
                    13.42,
                    -25.43,
                    '',
                    '',
                    'takeaway sit_down',
                    1,
                    1,
                ],
                [
                    'Harvest',
                    '12.43 -24.53',
                    12.43,
                    -24.53,
                    '',
                    '',
                    'sit_down',
                    1,
                    0,
                ],
                ['Wololo', '12.43 -24.54 1 0', 12.43, -24.54, 1, 0, '', 0, 0],
                [
                    'Los pollos hermanos',
                    '12.43 -24.54 1',
                    12.43,
                    -24.54,
                    1,
                    '',
                    '',
                    '',
                    '',
                ],
            ],
        }

        self.assertEqual(export, expected)

    def test_xlsx(self):
        title, schemas, submissions = build_fixture('grouped_repeatable')
        fp = FormPack(schemas, title)
        options = {'versions': 'rgv1'}

        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export(**options).to_xlsx(xls, submissions)
            assert xls.isfile()

    def test_xlsx_with_types(self):
        title, schemas, submissions = build_fixture('nested_grouped_repeatable')
        fp = FormPack(schemas, title)
        options = {'versions': 'bird_nests_v2', 'xls_types_as_text': False}

        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export(**options).to_xlsx(xls, submissions)
            assert xls.isfile()

    def test_xlsx_long_sheet_names_and_invalid_chars(self):
        title, schemas, submissions = build_fixture('long_names')
        fp = FormPack(schemas, title)
        options = {
            'versions': 'long_survey_name__the_quick__brown_fox_jumps'
            '_over_the_lazy_dog_v1'
        }

        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export(**options).to_xlsx(xls, submissions)
            assert xls.isfile()
            book = openpyxl.load_workbook(xls)
            assert book.sheetnames == [
                'long survey name_ the quick,...',
                'long_group_name__Victor_jagt...',
                'long_group_name__Victor_... (1)',
            ]

    def test_xlsx_too_long_string(self):
        """
        Make sure a warning is prepended when a response exceeds the maximum
        number of characters allowed in an Excel cell. Also, verify that
        subsequent responses in the same row render properly (see #309).
        """
        MAX_CHARS_IN_CELL = 32767
        title, schemas, submissions = build_fixture(
            'quotes_newlines_and_long_urls'
        )
        fp = FormPack(schemas, title)
        too_long = 'x' * (MAX_CHARS_IN_CELL + 1)
        submissions[0]['Enter_some_long_text_and_linebreaks_here'] = too_long
        expected_first_row = [
            (
                '<WARNING: Truncated to Excel limit of 32767 characters!>'
                + too_long
            )[:MAX_CHARS_IN_CELL],
            'yes',
        ]
        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export().to_xlsx(xls, submissions)
            assert xls.isfile()
            book = openpyxl.load_workbook(xls)
            sheet = book[title]
            row_values = [cell.value for cell in sheet[2]]
            assert row_values == expected_first_row

    def test_xlsx_too_long_url(self):
        """
        A too-long URL should be written, intact, as a plain string.
        """
        title, schemas, submissions = build_fixture(
            'quotes_newlines_and_long_urls'
        )
        fp = FormPack(schemas, title)
        expected_first_row = [
            'Check out this URL I found:\nhttps://now.read.this/?Never%20forg'
            'et%20that%20you%20are%20one%20of%20a%20kind.%20Never%20forget%20t'
            'hat%20if%20there%20weren%27t%20any%20need%20for%20you%20in%20all%'
            '20your%20uniqueness%20to%20be%20on%20this%20earth%2C%20you%20woul'
            'dn%27t%20be%20here%20in%20the%20first%20place.%20And%20never%20fo'
            'rget%2C%20no%20matter%20how%20overwhelming%20life%27s%20challenge'
            's%20and%20problems%20seem%20to%20be%2C%20that%20one%20person%20ca'
            'n%20make%20a%20difference%20in%20the%20world.%20In%20fact%2C%20it'
            '%20is%20always%20because%20of%20one%20person%20that%20all%20the%2'
            '0changes%20that%20matter%20in%20the%20world%20come%20about.%20So%'
            '20be%20that%20one%20person' + ('!' * 3000),
            'yes'
        ]
        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export().to_xlsx(xls, submissions)
            assert xls.isfile()
            book = openpyxl.load_workbook(xls)
            sheet = book[title]
            row_values = [cell.value for cell in sheet[2]]
            assert row_values == expected_first_row

    @pytest.mark.slow
    def test_xlsx_too_many_urls(self):
        """
        Excel doesn't allow more than 65,530 URLs per worksheet. URLs beyond
        this limit should be written as plain strings.
        """
        GOODNESS_LOOK_AT_THOSE_URLS = 65530
        title, schemas, submissions = build_fixture(
            'quotes_newlines_and_long_urls'
        )
        version = submissions[0]['__version__']
        # There's already a URL in the fixture data, but add 1 here just to
        # make it obvious that the limit is being exceeded
        for i in range(GOODNESS_LOOK_AT_THOSE_URLS + 1):
            submissions.append(
                {
                    'Enter_some_long_text_and_linebreaks_here': f'http://{i}',
                    'Some_other_question': 'yes',
                    '__version__': version,
                }
            )
        expected_last_row = [f'http://{i}', 'yes']
        fp = FormPack(schemas, title)
        # Faster than writing to a file, but still takes about 5 seconds
        temporary_xlsx = io.BytesIO()
        fp.export().to_xlsx(temporary_xlsx, submissions)
        book = openpyxl.load_workbook(temporary_xlsx, read_only=True)
        sheet = book[title]
        row_values = [cell.value for cell in sheet[len(submissions) + 1]]
        assert row_values == expected_last_row

    def test_xlsx_with_tag_headers(self):
        title, schemas, submissions = build_fixture('hxl_grouped_repeatable')
        fp = FormPack(schemas, title)
        options = {'versions': 'hxl_rgv1', 'tag_cols_for_header': ['hxl']}
        with TempDir() as d:
            xls = d / 'foo.xlsx'
            fp.export(**options).to_xlsx(xls, submissions)
            assert xls.isfile()
            book = openpyxl.load_workbook(xls, data_only=True)
            # Verify main sheet
            sheet = book['Household survey with HXL an...']
            row_values = [cell.value for cell in sheet[2]]
            assert row_values == ['#date+start', '#date+end', '#loc+name', None]
            # Verify repeating group
            sheet = book['houshold_member_repeat']
            row_values = [cell.value for cell in sheet[2]]
            assert row_values == ['#beneficiary', None, None]

    def test_force_index(self):
        title, schemas, submissions = customer_satisfaction

        forms = FormPack(schemas, title)
        export = forms.export(force_index=True).to_dict(submissions)
        expected = OrderedDict(
            {
                'Customer Satisfaction': {
                    'fields': [
                        'restaurant_name',
                        'customer_enjoyment',
                        '_index',
                    ],
                    'data': [
                        ['Felipes', 'yes', 1],
                        ['Dunkin Donuts', 'no', 2],
                        ['McDonalds', 'no', 3],
                    ],
                }
            }
        )

        self.assertEqual(export, expected)

    def test_copy_fields(self):
        title, schemas, submissions = customer_satisfaction

        forms = FormPack(schemas, title)
        export = forms.export(
            copy_fields=('_uuid', '_submission_time', ValidationStatusCopyField)
        )
        exported = export.to_dict(submissions)
        expected = OrderedDict(
            {
                'Customer Satisfaction': {
                    'fields': [
                        'restaurant_name',
                        'customer_enjoyment',
                        '_uuid',
                        '_submission_time',
                        '_validation_status',
                    ],
                    'data': [
                        [
                            'Felipes',
                            'yes',
                            '90dd7750f83011e590707c7a9125d07d',
                            '2016-04-01 19:57:45.306805',
                            'validation_status_approved',
                        ],
                        [
                            'Dunkin Donuts',
                            'no',
                            '90dd7750f83011e590707c7a9125d08d',
                            '2016-04-02 19:57:45.306805',
                            'validation_status_approved',
                        ],
                        [
                            'McDonalds',
                            'no',
                            '90dd7750f83011e590707c7a9125d09d',
                            '2016-04-03 19:57:45.306805',
                            'validation_status_approved',
                        ],
                    ],
                }
            }
        )

        self.assertDictEquals(exported, expected)

    def test_copy_fields_and_force_index_and_unicode(self):
        title, schemas, submissions = customer_satisfaction

        fp = FormPack(schemas, 'رضا العملاء')
        export = fp.export(
            copy_fields=(
                '_uuid',
                '_submission_time',
                ValidationStatusCopyField,
            ),
            force_index=True,
        )
        exported = export.to_dict(submissions)
        expected = OrderedDict(
            {
                'رضا العملاء': {
                    'fields': [
                        'restaurant_name',
                        'customer_enjoyment',
                        '_uuid',
                        '_submission_time',
                        '_validation_status',
                        '_index',
                    ],
                    'data': [
                        [
                            'Felipes',
                            'yes',
                            '90dd7750f83011e590707c7a9125d07d',
                            '2016-04-01 19:57:45.306805',
                            'validation_status_approved',
                            1,
                        ],
                        [
                            'Dunkin Donuts',
                            'no',
                            '90dd7750f83011e590707c7a9125d08d',
                            '2016-04-02 19:57:45.306805',
                            'validation_status_approved',
                            2,
                        ],
                        [
                            'McDonalds',
                            'no',
                            '90dd7750f83011e590707c7a9125d09d',
                            '2016-04-03 19:57:45.306805',
                            'validation_status_approved',
                            3,
                        ],
                    ],
                }
            }
        )

        self.assertEqual(exported, expected)

        with TempDir() as d:
            xls = d / 'test.xlsx'
            fp.export().to_xlsx(xls, submissions)
            assert xls.isfile()

    def test_copy_fields_multiple_versions(self):
        title, schemas, submissions = restaurant_profile

        forms = FormPack(schemas, title)
        export = forms.export(versions=forms.versions, copy_fields=('_uuid',))
        exported = export.to_dict(submissions)
        expected = OrderedDict(
            {
                'Restaurant profile': {
                    'fields': [
                        'restaurant_name',
                        'location',
                        '_location_latitude',
                        '_location_longitude',
                        '_location_altitude',
                        '_location_precision',
                        'eatery_type',
                        'eatery_type/sit_down',
                        'eatery_type/takeaway',
                        '_uuid',
                    ],
                    'data': [
                        [
                            'Felipes',
                            '12.34 -23.45',
                            '12.34',
                            '-23.45',
                            '',
                            '',
                            '',
                            '',
                            '',
                            '5dd6ecda-b993-42fc-95c2-7856a8940acf',
                        ],
                        [
                            'Felipes',
                            '12.34 -23.45',
                            '12.34',
                            '-23.45',
                            '',
                            '',
                            '',
                            '',
                            '',
                            'd6dee2e1-e0e6-4d08-9ad4-d78d77079f85',
                        ],
                        [
                            'Taco Truck',
                            '13.42 -25.43',
                            '13.42',
                            '-25.43',
                            '',
                            '',
                            'takeaway',
                            '',
                            '',
                            '3f2ac742-305a-4b0d-b7ef-f7f57fcd14dc',
                        ],
                        [
                            'Harvest',
                            '12.43 -24.53',
                            '12.43',
                            '-24.53',
                            '',
                            '',
                            'sit_down',
                            '',
                            '',
                            '3195b926-1578-4bac-80fc-735129a34090',
                        ],
                        [
                            'Taco Truck',
                            '13.42 -25.43',
                            '13.42',
                            '-25.43',
                            '',
                            '',
                            'takeaway sit_down',
                            '1',
                            '1',
                            '04cbcf32-ecbd-4801-829b-299463dcd125',
                        ],
                        [
                            'Harvest',
                            '12.43 -24.53',
                            '12.43',
                            '-24.53',
                            '',
                            '',
                            'sit_down',
                            '1',
                            '0',
                            '1f21b881-db1d-4629-9b82-f4111630187d',
                        ],
                        [
                            'Wololo',
                            '12.43 -24.54 1 0',
                            '12.43',
                            '-24.54',
                            '1',
                            '0',
                            '',
                            '0',
                            '0',
                            'fda7e49b-6c84-4cfe-b1a8-3de997ac0880',
                        ],
                        [
                            'Los pollos hermanos',
                            '12.43 -24.54 1',
                            '12.43',
                            '-24.54',
                            '1',
                            '',
                            '',
                            '',
                            '',
                            'a4277940-c8f3-4564-ad3b-14e28532a976',
                        ],
                    ],
                }
            }
        )

        self.assertDictEquals(exported, expected)

    def test_choices_external_as_text_field(self):
        title, schemas, submissions = build_fixture(
            'sanitation_report_external'
        )

        fp = FormPack(schemas, title)
        export = fp.export(lang=UNTRANSLATED)
        exported = export.to_dict(submissions)
        expected = OrderedDict(
            [
                (
                    'Sanitation report external',
                    {
                        'fields': [
                            'Restaurant name',
                            'How did this restaurant do on its sanitation report?',
                            'Report date',
                        ],
                        'data': [
                            ['Felipes', 'A', '012345'],
                            ['Chipotle', 'C', '012346'],
                            ['Dunkin Donuts', 'D', '012347'],
                            ['Boloco', 'B', '012348'],
                        ],
                    },
                )
            ]
        )

        self.assertEqual(exported, expected)

    def test_headers_of_multi_version_exports(self):
        title, schemas, submissions = build_fixture('site_inspection')
        fp = FormPack(schemas, title)
        export = fp.export(versions=fp.versions.keys()).to_dict(submissions)
        headers = export['Site inspection']['fields']
        self.assertListEqual(
            headers,
            [
                'inspector',
                'did_you_find_the_site',
                'was_there_damage_to_the_site',
                'ping',
                'rssi',
                'is_the_gate_secure',
                'is_plant_life_encroaching',
                'please_rate_the_impact_of_any_defects_observed',
                'was_there_damage_to_the_site_dupe',
            ],
        )

    def test_literacy_test_export(self):
        title, schemas, submissions = build_fixture('literacy_test')
        fp = FormPack(schemas, title)
        export = fp.export(versions=fp.versions.keys()).to_dict(submissions)
        headers = export['Literacy test']['fields']
        expected_headers = (
            [
                'russian_passage_1/Word at flash',
                'russian_passage_1/Duration of exercise',
                'russian_passage_1/Total words attempted',
                'russian_passage_1',
            ]
            + ['russian_passage_1/' + str(i) for i in range(1, 47)]
            + [
                'russian_passage_2/Word at flash',
                'russian_passage_2/Duration of exercise',
                'russian_passage_2/Total words attempted',
                'russian_passage_2',
            ]
            + ['russian_passage_2/' + str(i) for i in range(1, 47)]
        )
        self.assertListEqual(headers, expected_headers)
        # avoid heinous python-black reformatting where
        # EVERY 👏 SINGLE 👏 ITEM 👏 gets put on a separate line.
        # fmt: off
        expected_data = [
            [
                # Word at flash, duration of exercise, total words attempted
                '22', '16', '46',
                # Incorrect words as list in single column
                '1 2 4 8 10 19 20 21 29 30 33 39 46',
                # Incorrect words as binary values in separate columns
                '1', '1', '0', '1', '0', '0', '0', '1', '0', '1', '0', '0',
                '0', '0', '0', '0', '0', '0', '1', '1', '1', '0', '0', '0',
                '0', '0', '0', '0', '1', '1', '0', '0', '1', '0', '0', '0',
                '0', '0', '1', '0', '0', '0', '0', '0', '0', '1',
                # All the same for the  second literacy test question
                '21', '9', '46',
                '1 5 14 16 30 32 39',
                '1', '0', '0', '0', '1', '0', '0', '0', '0', '0', '0', '0',
                '0', '1', '0', '1', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '1', '0', '1', '0', '0', '0', '0',
                '0', '0', '1', '0', '0', '0', '0', '0', '0', '0',
            ], [
                '46', '14', '46',
                '1 2 3 4 5 6 7 8 9',
                '1', '1', '1', '1', '1', '1', '1', '1', '1', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '45', '7', '46',
                '1 11 29 46',
                '1', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '1', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '1',
            ], [
                '9', '12', '46',
                '1 2 3 4 6 7 8',
                '1', '1', '1', '1', '0', '1', '1', '1', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
                '33', '7', '36',
                '1 11 20 30 32',
                '1', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1', '0',
                '0', '0', '0', '0', '0', '0', '0', '1', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '1', '0', '1', '0', '0', '0', '0',
                '0', '0', '0', '0', '0', '0', '0', '0', '0', '0',
            ],
        ]
        # fmt: on
        self.assertListEqual(export['Literacy test']['data'], expected_data)

    def test_headers_of_multi_version_exports_with_copy_fields(self):
        title, schemas, submissions = build_fixture(
            'site_inspection', 'DATA_WITH_COPY_FIELDS'
        )

        fp = FormPack(schemas, title)
        export = fp.export(
            versions=fp.versions.keys(),
            copy_fields=(
                '_id',
                '_uuid',
                '_submission_time',
                ValidationStatusCopyField,
            ),
        ).to_dict(submissions)
        headers = export['Site inspection']['fields'][-4:]
        self.assertListEqual(
            headers, ['_id', '_uuid', '_submission_time', '_validation_status']
        )

    def test_spss_labels(self):
        fixture_name = 'long_unicode_labels'
        title, schemas, submissions = build_fixture(fixture_name)
        fp = FormPack(schemas, title)
        options = {
            'versions': 'long_unicode_labels_v1',
        }
        expected_label_file_names = [
            'long unicode labels to test SPSS export - English - SPSS labels.sps',
            'long unicode labels to test SPSS export - Français - SPSS labels.sps',
            'long unicode labels to test SPSS export - Swahili - SPSS labels.sps',
        ]
        # Export to an in-memory ZIP file
        raw_zip = BytesIO()
        fp.export(**options).to_spss_labels(raw_zip)
        raw_zip.seek(0)
        zipped = ZipFile(raw_zip, 'r')
        for name in expected_label_file_names:
            with open_fixture_file(fixture_name, name, 'r') as expected:
                actual = TextIOWrapper(
                    zipped.open(name, 'r'), newline=None, encoding='utf-8'
                )
                actual_content = actual.read()
                expected_content = expected.read()

                # ToDo remove condition when Python2 support is dropped
                if isinstance(actual_content, bytes):
                    actual_content = actual_content.decode('utf-8')

                assert actual_content == expected_content
        zipped.close()
        raw_zip.close()

    def test_untranslated_spss_labels(self):
        fixture_name = 'long_unicode_labels'
        title, schemas, submissions = build_fixture(fixture_name)
        # Remove every language except the first
        content = schemas[0]['content']
        first_translation = content['translations'][0]
        for sheet in 'survey', 'choices':
            for row in content[sheet]:
                for col in content['translated']:
                    try:
                        # Replace list of translations with first translation
                        row[col] = row[col][0]
                    except KeyError:
                        pass
        content['translated'] = []
        content['translations'] = [None]
        # Proceed with the export
        fp = FormPack(schemas, title)
        options = {
            'versions': 'long_unicode_labels_v1',
        }
        fixture_label_file_name = (
            'long unicode labels to test SPSS export - {} - '
            'SPSS labels.sps'.format(first_translation)
        )
        expected_label_file_name = (
            'long unicode labels to test SPSS export - ' 'SPSS labels.sps'
        )
        # Export to an in-memory ZIP file
        raw_zip = BytesIO()
        fp.export(**options).to_spss_labels(raw_zip)
        raw_zip.seek(0)
        zipped = ZipFile(raw_zip, 'r')
        with open_fixture_file(
            fixture_name, fixture_label_file_name, 'r'
        ) as expected:
            actual = TextIOWrapper(
                zipped.open(expected_label_file_name, 'r'),
                newline=None,
                encoding='utf-8',
            )
            actual_content = actual.read()
            expected_content = expected.read()
            # ToDo remove condition when Python2 support is dropped
            if isinstance(actual_content, bytes):
                actual_content = actual_content.decode('utf-8')

            assert actual_content == expected_content
        zipped.close()
        raw_zip.close()

    def test_select_multiple_summary(self):
        title, schemas, submissions = build_fixture('dietary_needs')
        fp = FormPack(schemas, title)
        export = fp.export(
            multiple_select='summary', versions=fp.versions.keys()
        ).to_dict(submissions)
        expected = OrderedDict(
            [
                (
                    'Dietary needs',
                    {
                        'fields': ['restaurant_name', 'dietary_accommodations'],
                        'data': [
                            ["Melba's", 'gluten_free'],
                            ['Land of Kush', 'vegan vegetarian'],
                            [
                                'Sweet 27',
                                'gluten_free vegan vegetarian lactose_free',
                            ],
                        ],
                    },
                )
            ]
        )
        assert export == expected

    def test_select_multiple_details(self):
        title, schemas, submissions = build_fixture('dietary_needs')
        fp = FormPack(schemas, title)
        export = fp.export(
            multiple_select='details', versions=fp.versions.keys()
        ).to_dict(submissions)
        expected = OrderedDict(
            [
                (
                    'Dietary needs',
                    {
                        'fields': [
                            'restaurant_name',
                            'dietary_accommodations/gluten_free',
                            'dietary_accommodations/vegan',
                            'dietary_accommodations/vegetarian',
                            'dietary_accommodations/lactose_free',
                        ],
                        'data': [
                            ["Melba's", '1', '0', '0', '0'],
                            ['Land of Kush', '0', '1', '1', '0'],
                            ['Sweet 27', '1', '1', '1', '1'],
                        ],
                    },
                )
            ]
        )
        assert export == expected

    def test_select_multiple_both(self):
        title, schemas, submissions = build_fixture('dietary_needs')
        fp = FormPack(schemas, title)
        export = fp.export(
            multiple_select='both', versions=fp.versions.keys()
        ).to_dict(submissions)
        expected = OrderedDict(
            [
                (
                    'Dietary needs',
                    {
                        'fields': [
                            'restaurant_name',
                            'dietary_accommodations',
                            'dietary_accommodations/gluten_free',
                            'dietary_accommodations/vegan',
                            'dietary_accommodations/vegetarian',
                            'dietary_accommodations/lactose_free',
                        ],
                        'data': [
                            ["Melba's", 'gluten_free', '1', '0', '0', '0'],
                            [
                                'Land of Kush',
                                'vegan vegetarian',
                                '0',
                                '1',
                                '1',
                                '0',
                            ],
                            [
                                'Sweet 27',
                                'gluten_free vegan vegetarian lactose_free',
                                '1',
                                '1',
                                '1',
                                '1',
                            ],
                        ],
                    },
                )
            ]
        )
        assert export == expected

    def test_select_multiple_with_different_options_in_multiple_versions(self):
        title, schemas, submissions = build_fixture('favorite_coffee')
        fp = FormPack(schemas, title)
        self.assertEqual(len(fp.versions), 2)

        export = fp.export(versions=fp.versions.keys()).to_dict(submissions)

        headers = export['Favorite coffee']['fields']
        self.assertListEqual(
            headers,
            [
                'favorite_coffee_type',
                'favorite_coffee_type/french',
                'favorite_coffee_type/italian',
                'favorite_coffee_type/american',
                'favorite_coffee_type/british',
                'brand_of_coffee_machine',
            ],
        )

        # Check length of each row
        for row in export['Favorite coffee']['data']:
            self.assertEqual(len(headers), len(row))

        # Ensure latest submissions is not shifted
        self.assertListEqual(
            export['Favorite coffee']['data'][-1],
            ['american british', '0', '0', '1', '1', 'Keurig'],
        )

    def test_geojson_with_select_xml_label(self):
        title, schemas, submissions = build_fixture('geojson_and_selects')
        fp = FormPack(schemas, title)

        options = {'versions': 'v1', 'lang': '_xml', 'include_media_url': True}
        export = fp.export(**options)
        geojson_gen = export.to_geojson(
            submissions, geo_question_name='geo_location'
        )
        geojson_str = ''.join(geojson_gen)
        geojson_obj = json.loads(geojson_str)

        assert geojson_obj == {
            'type': 'FeatureCollection',
            'name': 'Geo and selects',
            'features': [
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Point',
                        'coordinates': [
                            -76.60869,
                            39.306938,
                            11.0,
                        ],
                    },
                    'properties': {
                        'an_image': 'location.jpeg',
                        'an_image_URL': 'https://kc.kobo.org/media/original?media_file=/path/to/location.jpeg',
                        'current_location': 'inside',
                    },
                },
            ],
        }

    def test_geojson_with_select_label(self):
        title, schemas, submissions = build_fixture('geojson_and_selects')
        fp = FormPack(schemas, title)

        options = {
            'versions': 'v1',
            'lang': UNTRANSLATED,
            'include_media_url': True,
        }
        export = fp.export(**options)
        geojson_gen = export.to_geojson(
            submissions, geo_question_name='geo_location'
        )
        geojson_str = ''.join(geojson_gen)
        geojson_obj = json.loads(geojson_str)

        assert geojson_obj == {
            'type': 'FeatureCollection',
            'name': 'Geo and selects',
            'features': [
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Point',
                        'coordinates': [
                            -76.60869,
                            39.306938,
                            11.0,
                        ],
                    },
                    'properties': {
                        'Take a photo of the location': 'location.jpeg',
                        'Take a photo of the location_URL': 'https://kc.kobo.org/media/original?media_file=/path/to/location.jpeg',
                        'Where are you?': 'Inside',
                    },
                },
            ],
        }

    def test_geojson_point(self):
        title, schemas, submissions = build_fixture('all_geo_types')
        fp = FormPack(schemas, title)
        assert len(fp.versions) == 2

        export = fp.export(versions=fp.versions.keys())
        geojson_gen = export.to_geojson(submissions, geo_question_name='Point')
        geojson_str = ''.join(geojson_gen)
        geojson_obj = json.loads(geojson_str)

        assert geojson_obj == {
            'type': 'FeatureCollection',
            'name': 'I have points, traces, and shapes!',
            'features': [
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Point',
                        'coordinates': [
                            -76.60869,
                            39.306938,
                            11.0,
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:42:37.313-04:00',
                        'end': '2019-07-19T18:47:10.516-04:00',
                        'Just_a_regular_text_question': 'Greenmount',
                    },
                },
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Point',
                        'coordinates': [
                            -58.442238,
                            -34.631098,
                            0.0,
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:49:05.982-04:00',
                        'end': '2019-07-19T19:03:07.602-04:00',
                        'Just_a_regular_text_question': 'Chacabuco',
                    },
                },
            ],
        }

    def test_geojson_trace(self):
        title, schemas, submissions = build_fixture('all_geo_types')
        fp = FormPack(schemas, title)
        assert len(fp.versions) == 2

        export = fp.export(versions=fp.versions.keys())
        geojson_gen = export.to_geojson(submissions, geo_question_name='Trace')
        geojson_str = ''.join(geojson_gen)
        geojson_obj = json.loads(geojson_str)

        assert geojson_obj == {
            'type': 'FeatureCollection',
            'name': 'I have points, traces, and shapes!',
            'features': [
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'LineString',
                        'coordinates': [
                            [-76.608323, 39.306032, 1.0],
                            [-76.608953, 39.308701, 2.0],
                            [-76.60938, 39.311313, 3.0],
                            [-76.604223, 39.3116, 4.0],
                            [-76.603804, 39.306077, 5.0],
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:42:37.313-04:00',
                        'end': '2019-07-19T18:47:10.516-04:00',
                        'Just_a_regular_text_question': 'Greenmount',
                    },
                },
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'LineString',
                        'coordinates': [
                            [-58.44218, -34.631089, 0.0],
                            [-58.439112, -34.635079, 0.0],
                            [-58.445635, -34.636562, 0.0],
                            [-58.44713, -34.634929, 0.0],
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:49:05.982-04:00',
                        'end': '2019-07-19T19:03:07.602-04:00',
                        'Just_a_regular_text_question': 'Chacabuco',
                    },
                },
            ],
        }

    def test_geojson_shape(self):
        title, schemas, submissions = build_fixture('all_geo_types')
        fp = FormPack(schemas, title)
        assert len(fp.versions) == 2

        export = fp.export(versions=fp.versions.keys())
        geojson_gen = export.to_geojson(submissions, geo_question_name='Shape')
        geojson_str = ''.join(geojson_gen)
        geojson_obj = json.loads(geojson_str)

        assert geojson_obj == {
            'type': 'FeatureCollection',
            'name': 'I have points, traces, and shapes!',
            'features': [
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Polygon',
                        'coordinates': [
                            [
                                [-76.608294, 39.305938, 0],
                                [-76.603656, 39.306138, 0],
                                [-76.604171, 39.31155, 0],
                                [-76.609332, 39.311288, 0],
                                [-76.609211, 39.309365, 0],
                                [-76.608294, 39.305938, 0],
                            ]
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:42:37.313-04:00',
                        'end': '2019-07-19T18:47:10.516-04:00',
                        'Just_a_regular_text_question': 'Greenmount',
                    },
                },
                {
                    'type': 'Feature',
                    'geometry': {
                        'type': 'Polygon',
                        'coordinates': [
                            [
                                [-58.447229, -34.6347, 0],
                                [-58.445613, -34.636607, 0],
                                [-58.438946, -34.635194, 0],
                                [-58.442056, -34.631063, 0],
                                [-58.447229, -34.6347, 0],
                            ]
                        ],
                    },
                    'properties': {
                        'start': '2019-07-19T18:49:05.982-04:00',
                        'end': '2019-07-19T19:03:07.602-04:00',
                        'Just_a_regular_text_question': 'Chacabuco',
                    },
                },
            ],
        }

    def test_geojson_invalid(self):
        title, schemas, _ = build_fixture('all_geo_types')
        fp = FormPack(schemas, title)
        self.assertEqual(len(fp.versions), 2)

        # Can't test an invalid `geopoint` by itself; it'll blow up
        # `FormGPSField.format()`
        submissions = [
            {'Trace': '1 2 3 4'},  # Not enough points
            {'Trace': '1 2 3 4;1'},  # Second point is bogus
            {'Trace': '1 2 3 4;1 2 banana'},  # Second point is still bogus
            {'Shape': '1 2 3 4;5 6 7 8;9 10 11 12;13 14 15 16'},  # Not closed
            # The following are okay and must appear in the export
            {'Trace': '1 2 3 4;5 6 7 8'},
            {'Shape': '1 2 3 4;5 6 7 8;9 10 11 12;1 2 3 4'},
        ]
        for s in submissions:
            s[fp.default_version_id_key] = get_first_occurrence(fp.versions)
        export = fp.export(versions=fp.versions.keys())
        geojson_obj = json.loads(
            ''.join(export.to_geojson(submissions, flatten=True))
        )
        assert len(geojson_obj['features']) == 2
        assert geojson_obj['features'][0]['geometry'] == {
            'coordinates': [[2.0, 1.0, 3.0], [6.0, 5.0, 7.0]],
            'type': 'LineString',
        }
        assert geojson_obj['features'][1]['geometry'] == {
            'coordinates': [
                [
                    [2.0, 1.0, 3.0],
                    [10.0, 9.0, 11.0],
                    [6.0, 5.0, 7.0],
                    [2.0, 1.0, 3.0],
                ]
            ],
            'type': 'Polygon',
        }

    def test_geojson_unflattened(self):
        title, schemas, submissions = build_fixture('all_geo_types')
        fp = FormPack(schemas, title)
        assert len(fp.versions) == 2

        export = fp.export(versions=fp.versions.keys())
        geojson_gen = export.to_geojson(submissions, flatten=False)
        geojson_str = ''.join(geojson_gen)

        geojson_obj = json.loads(geojson_str)
        assert len(geojson_obj) == 2
        assert len(geojson_obj[0]['features']) == 3
        geojson_obj_f_types = [
            ft['geometry']['type'] for ft in geojson_obj[0]['features']
        ]
        for f_type in ('Point', 'LineString', 'Polygon'):
            assert f_type in geojson_obj_f_types

        assert geojson_obj == [
            {
                'type': 'FeatureCollection',
                'name': 'I have points, traces, and shapes!',
                'features': [
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'Point',
                            'coordinates': [-76.60869, 39.306938, 11.0],
                        },
                        'properties': {
                            'start': '2019-07-19T18:42:37.313-04:00',
                            'end': '2019-07-19T18:47:10.516-04:00',
                            'Just_a_regular_text_question': 'Greenmount',
                        },
                    },
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'LineString',
                            'coordinates': [
                                [-76.608323, 39.306032, 1.0],
                                [-76.608953, 39.308701, 2.0],
                                [-76.60938, 39.311313, 3.0],
                                [-76.604223, 39.3116, 4.0],
                                [-76.603804, 39.306077, 5.0],
                            ],
                        },
                        'properties': {
                            'start': '2019-07-19T18:42:37.313-04:00',
                            'end': '2019-07-19T18:47:10.516-04:00',
                            'Just_a_regular_text_question': 'Greenmount',
                        },
                    },
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'Polygon',
                            'coordinates': [
                                [
                                    [-76.608294, 39.305938, 0.0],
                                    [-76.603656, 39.306138, 0.0],
                                    [-76.604171, 39.31155, 0.0],
                                    [-76.609332, 39.311288, 0.0],
                                    [-76.609211, 39.309365, 0.0],
                                    [-76.608294, 39.305938, 0.0],
                                ]
                            ],
                        },
                        'properties': {
                            'start': '2019-07-19T18:42:37.313-04:00',
                            'end': '2019-07-19T18:47:10.516-04:00',
                            'Just_a_regular_text_question': 'Greenmount',
                        },
                    },
                ],
            },
            {
                'type': 'FeatureCollection',
                'name': 'I have points, traces, and shapes!',
                'features': [
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'Point',
                            'coordinates': [-58.442238, -34.631098, 0.0],
                        },
                        'properties': {
                            'start': '2019-07-19T18:49:05.982-04:00',
                            'end': '2019-07-19T19:03:07.602-04:00',
                            'Just_a_regular_text_question': 'Chacabuco',
                        },
                    },
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'LineString',
                            'coordinates': [
                                [-58.44218, -34.631089, 0.0],
                                [-58.439112, -34.635079, 0.0],
                                [-58.445635, -34.636562, 0.0],
                                [-58.44713, -34.634929, 0.0],
                            ],
                        },
                        'properties': {
                            'start': '2019-07-19T18:49:05.982-04:00',
                            'end': '2019-07-19T19:03:07.602-04:00',
                            'Just_a_regular_text_question': 'Chacabuco',
                        },
                    },
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'Polygon',
                            'coordinates': [
                                [
                                    [-58.447229, -34.6347, 0.0],
                                    [-58.445613, -34.636607, 0.0],
                                    [-58.438946, -34.635194, 0.0],
                                    [-58.442056, -34.631063, 0.0],
                                    [-58.447229, -34.6347, 0.0],
                                ]
                            ],
                        },
                        'properties': {
                            'start': '2019-07-19T18:49:05.982-04:00',
                            'end': '2019-07-19T19:03:07.602-04:00',
                            'Just_a_regular_text_question': 'Chacabuco',
                        },
                    },
                    {
                        'type': 'Feature',
                        'geometry': {
                            'type': 'Point',
                            'coordinates': [-58.355287, -34.619206, 0.0],
                        },
                        'properties': {
                            'start': '2019-07-19T18:49:05.982-04:00',
                            'end': '2019-07-19T19:03:07.602-04:00',
                            'Just_a_regular_text_question': 'Chacabuco',
                        },
                    },
                ],
            },
        ]

    #https://github.com/kobotoolbox/formpack/pull/215
    def test_header_label_list_label(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': None, 'lang' : None}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["Restaurant name",
                           "Did you enjoy your dining experience?"],
                'data': [
                    ["Felipes", "Yes"],
                    ["Dunkin Donuts", "No"],
                    ["McDonalds", "No"]
                ]
            }
        })
        self.assertEqual(exported, expected)

    def test_header_key_list_key(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': False, 'lang' : False}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["restaurant_name",
                           "customer_enjoyment"],
                'data': [
                    ["Felipes", "yes"],
                    ["Dunkin Donuts", "no"],
                    ["McDonalds", "no"]
                ]
            }
        })
        self.assertEqual(exported, expected)

    def test_header_key_list_label(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': False, 'lang' : None}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["restaurant_name",
                           "customer_enjoyment"],
                'data': [
                    ["Felipes", "Yes"],
                    ["Dunkin Donuts", "No"],
                    ["McDonalds", "No"]
                ]
            }
        })
        self.assertEqual(exported, expected)

    def test_header_Label_list_key(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': None, 'lang' : False}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["Restaurant name",
                           "Did you enjoy your dining experience?"],
                'data': [
                    ["Felipes", "yes"],
                    ["Dunkin Donuts", "no"],
                    ["McDonalds", "no"]
                ]
            }
        })
        self.assertEqual(exported, expected)

    def test_header_label_no_lang(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': None}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["Restaurant name",
                           "Did you enjoy your dining experience?"],
                'data': [
                    ["Felipes", "yes"],
                    ["Dunkin Donuts", "no"],
                    ["McDonalds", "no"]
                ]
            }
        })
        self.assertEqual(exported, expected)

    def test_header_key_no_lang(self):
        title, schemas, submissions = customer_satisfaction
        fp = FormPack(schemas, title)
        options = {'header_lang': False}
        exported = fp.export(**options).to_dict(submissions)
        expected = OrderedDict({
            "Customer Satisfaction": {
                'fields': ["restaurant_name",
                           "customer_enjoyment"],
                'data': [
                    ["Felipes", "yes"],
                    ["Dunkin Donuts", "no"],
                    ["McDonalds", "no"]
                ]
            }
        })
        self.assertEqual(exported, expected)
