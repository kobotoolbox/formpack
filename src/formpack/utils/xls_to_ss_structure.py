# coding: utf-8
import datetime
import re
from collections import OrderedDict

import openpyxl
import xlrd

from .replace_aliases import kobo_specific_sub


def xls_to_lists(xls_file_object, strip_empty_rows=True):
    """
    The goal: Convert an XLS file object to a python object.

    This draws on code from `pyxform.xls2json_backends` and
    `convert_file_to_csv_string`, however this works as it is expected (does
    not add extra sheets or perform misc conversions which are a part of
    `pyxform.xls2json_backends.xls_to_dict`.)
    """

    def _iswhitespace(string):
        return isinstance(string, str) and len(string.strip()) == 0

    def xls_value_to_unicode(value, value_type):
        """
        Take a xls formatted value and try to make a unicode string
        representation.
        """
        if value_type == xlrd.XL_CELL_BOOLEAN:
            return 'TRUE' if value else 'FALSE'
        elif value_type == xlrd.XL_CELL_NUMBER:
            # Try to display as an int if possible.
            int_value = int(value)
            if int_value == value:
                return str(int_value)
            else:
                return str(value)
        elif value_type is xlrd.XL_CELL_DATE:
            # Warn that it is better to single quote as a string.
            # error_location = cellFormatString % (ss_row_idx, ss_col_idx)
            # raise Exception(
            #   "Cannot handle excel formatted date at " + error_location)
            datetime_or_time_only = xlrd.xldate_as_tuple(
                value, workbook.datemode
            )
            if datetime_or_time_only[:3] == (0, 0, 0):
                # must be time only
                return str(datetime.time(*datetime_or_time_only[3:]))
            return str(datetime.datetime(*datetime_or_time_only))
        else:
            # ensure unicode and replace nbsp spaces with normal ones
            # to avoid this issue:
            # https://github.com/modilabs/pyxform/issues/83
            return str(value).replace(chr(160), ' ')

    def _escape_newline_chars(cell):
        return re.sub(r'\r', '\\\\r', re.sub(r'\n', '\\\\n', cell))

    def _sheet_to_lists(sheet):
        result = []
        nrows_range = list(range(0, sheet.nrows))
        ncols_range = list(range(0, sheet.ncols))
        for row in nrows_range:
            row_results = []
            row_empty = True
            for col in ncols_range:
                value = sheet.cell_value(row, col)
                if isinstance(value, str):
                    value = _escape_newline_chars(value.strip())
                if (value is not None) and (not _iswhitespace(value)):
                    value = xls_value_to_unicode(
                        value, sheet.cell_type(row, col)
                    )
                if value != '':
                    row_empty = False
                if value == '':
                    value = None
                row_results.append(value)
            if not strip_empty_rows or not row_empty:
                result.append(row_results)
        return result

    workbook = xlrd.open_workbook(file_contents=xls_file_object.read())
    ss_structure = OrderedDict()
    for sheet in workbook.sheets():
        sheet_name = kobo_specific_sub(sheet.name)
        sheet_contents = _sheet_to_lists(sheet)
        ss_structure[sheet_name] = sheet_contents
    return ss_structure


def _parsed_sheet(sheet_lists):
    """
    take a sheet with 2+ rows and parse the first row as the column headers
    and the subsequent rows as the values.

    outputs a list of ordered dicts
    """
    # Treat sheets without at least two rows, i.e. without a header row
    # and at least one data row, as empty
    if len(sheet_lists) < 2:
        return []
    columns = sheet_lists[0]
    rows = sheet_lists[1:]
    out_list = []
    columns_range = list(range(0, len(columns)))
    for row in rows:
        out_row = OrderedDict()
        for ii in columns_range:
            if row[ii] is not None:
                out_row[columns[ii]] = row[ii]
        out_list.append(out_row)
    return out_list


def xls_to_dicts(xls_file_object, strip_empty_rows=True):
    """
    outputs an ordered dict of (sheetname, sheet_contents)

    where sheet_contents is a list of ordered_dicts
    """
    lists = xls_to_lists(xls_file_object)
    out = OrderedDict()
    for key, sheet in lists.items():
        out[key] = _parsed_sheet(sheet)
    return out


def xlsx_to_lists(xls_file_object, strip_empty_rows=True):
    """
    The goal: Convert an XLS file object to a python object.
    This draws on code from `pyxform.xls2json_backends` and
    `convert_file_to_csv_string`, however this works as it is expected (does
    not add extra sheets or perform misc conversions which are a part of
    `pyxform.xls2json_backends.xls_to_dict`.)
    """

    workbook = openpyxl.load_workbook(xls_file_object)

    def is_empty(value):
        if value is None:
            return True
        elif isinstance(value, str) and value.strip() == '':
            return True
        else:
            return False

    def xlsx_value_to_str(value):
        """
        Take a xls formatted value and try to make a string representation.
        """
        if value is True:
            return 'TRUE'
        elif value is False:
            return 'FALSE'
        elif isinstance(value, float) and value.is_integer():
            # Try to display as an int if possible.
            return str(int(value))
        elif isinstance(value, (int, datetime.datetime, datetime.time)):
            return str(value)
        else:
            # ensure unicode and replace nbsp spaces with normal ones
            # to avoid this issue:
            # https://github.com/modilabs/pyxform/issues/83
            return str(value).replace(chr(160), ' ')

    def xlsx_to_dict_normal_sheet(sheet):

        # Check for duplicate column headers
        column_header_list = list()
        for cell in sheet[1]:
            column_header = cell.value
            # xls file with 3 columns mostly have a 3 more columns that are
            # blank by default or something, skip during check
            if is_empty(column_header):
                # Preserve column order (will filter later)
                column_header_list.append(None)
            else:
                clean_header = re.sub(r'( )+', ' ', column_header.strip())
                column_header_list.append(clean_header)

        result = []
        for row in sheet.iter_rows(min_row=2):
            row_dict = OrderedDict()
            for column, key in enumerate(column_header_list):
                if key is None:
                    continue

                value = row[column].value
                if isinstance(value, str):
                    value = value.strip()

                if not is_empty(value):
                    row_dict[key] = xlsx_value_to_str(value)

            result.append(row_dict)

        return result

    result = OrderedDict()
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        sheetname = kobo_specific_sub(sheetname)
        result[sheetname] = xlsx_to_dict_normal_sheet(sheet)

    return result


def xlsx_to_dicts(xlsx_file_object, strip_empty_rows=True):
    """
    outputs an ordered dict of (sheetname, sheet_contents)
    where sheet_contents is a list of ordered_dicts
    """
    lists = xlsx_to_lists(xlsx_file_object)
    out = OrderedDict()
    for key, sheet in lists.items():
        out[key] = sheet
    return out
